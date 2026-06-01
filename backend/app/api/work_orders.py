from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session

from app.database import get_db
from app.schemas.work_order import (
    WorkOrderCreate,
    WorkOrderList,
    WorkOrderResponse,
    WorkOrderUpdate,
)
from app.services.work_order_service import WorkOrderService
from app.auth import require_admin
from app.models.user import User

router = APIRouter(prefix="/api/v1/work-orders", tags=["工单管理"])


@router.get("/", response_model=WorkOrderList)
def list_work_orders(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    business_city: Optional[str] = Query(None, description="业务受理地市"),
    product_category: Optional[str] = Query(None, description="产品分类"),
    operation_type: Optional[str] = Query(None, description="操作类型"),
    business_location_city: Optional[str] = Query(None, description="业务所属地市"),
    product_category_group: Optional[str] = Query(None, description="产品分类分组: 互联网专线/千里眼"),
    db: Session = Depends(get_db),
):
    items, total = WorkOrderService.list_work_orders(
        db,
        skip=skip,
        limit=limit,
        status=status,
        priority=priority,
        search=search,
        business_city=business_city,
        product_category=product_category,
        operation_type=operation_type,
        business_location_city=business_location_city,
        product_category_group=product_category_group,
    )
    return WorkOrderList(
        total=total,
        items=[WorkOrderResponse.model_validate(r) for r in items],
        skip=skip,
        limit=limit,
    )


@router.get("/cities", response_model=list[str])
def get_business_location_cities(
    db: Session = Depends(get_db),
):
    return WorkOrderService.get_business_location_cities(db)


@router.get("/{work_order_id}", response_model=WorkOrderResponse)
def get_work_order(work_order_id: int, db: Session = Depends(get_db)):
    wo = WorkOrderService.get_work_order(db, work_order_id)
    if not wo:
        raise HTTPException(status_code=404, detail="工单不存在")
    return WorkOrderResponse.model_validate(wo)


@router.post("/", response_model=WorkOrderResponse, status_code=201)
def create_work_order(
    data: WorkOrderCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    wo = WorkOrderService.create_work_order(db, data)
    return WorkOrderResponse.model_validate(wo)


@router.put("/{work_order_id}", response_model=WorkOrderResponse)
def update_work_order(
    work_order_id: int,
    data: WorkOrderUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    wo = WorkOrderService.update_work_order(db, work_order_id, data)
    if not wo:
        raise HTTPException(status_code=404, detail="工单不存在")
    return WorkOrderResponse.model_validate(wo)


import openpyxl

# 工单字段列表（与 Excel 列名对照）
WORK_ORDER_FIELDS = [
    "crm_order_id", "order_no", "title", "dispatch_time", "business_city",
    "product_category", "operation_type", "group_product_number",
    "business_location_city", "customer_address", "status", "current_step",
    "camera_install_location", "product_instance_id",
]

FIELD_LABELS = {
    "crm_order_id": "CRM订单号",
    "order_no": "工单号",
    "title": "工单主题",
    "dispatch_time": "派单时间",
    "business_city": "业务受理地市",
    "product_category": "产品分类",
    "operation_type": "操作类型",
    "group_product_number": "集团产品号码",
    "business_location_city": "业务所属地市",
    "customer_address": "客户机房详细地址",
    "status": "工单状态",
    "current_step": "当前环节名称",
    "camera_install_location": "摄像头安装位置",
    "product_instance_id": "产品实例编号",
}


@router.post("/upload", status_code=200)
def upload_work_orders(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """上传xlsx清单，遍历取消/开通/调整三个子表，完全替换工单数据。"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析Excel文件")

    # 需要读取的工作表名
    TARGET_SHEETS = ["取消", "开通", "调整"]
    all_items = []

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        if not rows or not headers:
            continue

        # 构建字段映射
        col_map = {}
        for idx, h in enumerate(headers):
            if h is None:
                continue
            h_clean = h.strip()
            for field, label in FIELD_LABELS.items():
                if h_clean == label:
                    col_map[idx] = field
                    break

        if not col_map:
            continue  # 跳过列名不匹配的工作表

        # 解析数据
        for row in rows:
            if all(v is None for v in row):
                continue
            item = {}
            for col_idx, field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    if hasattr(val, "strftime"):
                        val = val.strftime("%Y-%m-%d %H:%M")
                    else:
                        val = str(val).strip()
                item[field] = val
            all_items.append(item)

    if not all_items:
        raise HTTPException(
            status_code=400,
            detail="未找到有效数据，请确保Excel包含'取消'、'开通'、'调整'三个工作表，且列名与工单字段一致",
        )

    count = WorkOrderService.replace_all(db, all_items)
    return {"message": f"导入成功，共 {count} 条工单（取消/开通/调整）", "count": count}


@router.delete("/{work_order_id}", status_code=204)
def delete_work_order(
    work_order_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    ok = WorkOrderService.delete_work_order(db, work_order_id)
    if not ok:
        raise HTTPException(status_code=404, detail="工单不存在")
    return None
