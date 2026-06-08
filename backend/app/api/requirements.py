import csv
from io import StringIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.schemas.requirement import (
    RequirementCreate,
    RequirementList,
    RequirementResponse,
    RequirementUpdate,
)
from app.services.requirement_service import RequirementService, attach_req_status
from app.auth import require_admin
from app.models.user import User
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/requirements", tags=["需求管理"])

import openpyxl  # noqa: E402
import tempfile  # noqa: E402
import os  # noqa: E402

REQUIREMENT_FIELD_LABELS = {
    "receive_date": "接收日期",
    "owner_name": "业主姓名",
    "contact": "联系方式",
    "city": "市",
    "district": "区/县",
    "outlet_code": "网点号",
    "order_type": "工单类型",
    "install_address": "安装地址",
    "remark": "备注",
}


@router.post("/upload", status_code=200)
def upload_requirements(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """上传 xlsx/csv 清单导入需求，校验 网点号+工单类型+安装地址 唯一"""
    fn = (file.filename or "").lower()
    if not fn.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .csv 文件")

    is_csv = fn.endswith(".csv")
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    if is_csv:
        import io
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=400, detail="CSV 文件为空")
        headers = all_rows[0]
        rows = all_rows[1:]
        # Convert to openpyxl-like format: each row is a list of values
    else:
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            os.unlink(tmp_path)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"无法解析Excel文件: {str(e)}")

        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Excel 文件没有工作表")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if not rows or not headers:
        raise HTTPException(status_code=400, detail="文件无数据")

    col_map = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        h_clean = str(h).strip()
        for field, label in REQUIREMENT_FIELD_LABELS.items():
            if h_clean == label:
                col_map[idx] = field
                break

    if not col_map:
        raise HTTPException(
            status_code=400,
            detail="表头不匹配，请确保列名包含：" + "、".join(REQUIREMENT_FIELD_LABELS.values()),
        )

    all_items = []
    for row in rows:
        if all(v is None for v in row):
            continue
        item = {}
        for col_idx, field in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                if hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                else:
                    val = str(val).strip()
            item[field] = val
        all_items.append(item)

    if not all_items:
        raise HTTPException(status_code=400, detail="未找到有效数据")

    result = RequirementService.batch_import(db, all_items)
    return {
        "message": f"导入完成：成功 {result['imported']} 条，跳过 {result['skipped']} 条重复",
        **result,
    }


EXPORT_HEADERS = [
    ("receive_date", "接收日期"),
    ("req_status", "需求状态"),
    ("owner_name", "业主姓名"),
    ("contact", "联系方式"),
    ("city", "市"),
    ("district", "区/县"),
    ("outlet_code", "网点号"),
    ("order_type", "工单类型"),
    ("install_address", "安装地址"),
    ("remark", "备注"),
]


@router.get("/export")
def export_requirements(
    search: Optional[str] = Query(None),
    city: Optional[List[str]] = Query(None, description="市（多选）"),
    district: Optional[str] = Query(None, description="区/县"),
    order_type: Optional[List[str]] = Query(None, description="工单类型（多选）"),
    date_from: Optional[str] = Query(None, description="接收日期起"),
    date_to: Optional[str] = Query(None, description="接收日期止"),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """导出需求列表为 CSV（支持当前筛选条件）"""
    items, _ = RequirementService.list_requirements(
        db,
        skip=0,
        limit=100000,
        search=search,
        city=city,
        district=district,
        order_type=order_type,
        date_from=date_from,
        date_to=date_to,
    )

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([h[1] for h in EXPORT_HEADERS])

    for r in items:
        req_status = attach_req_status(r, db)
        writer.writerow([
            getattr(r, field, "")
            if field != "req_status" else req_status
            for field, _ in EXPORT_HEADERS
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=requirements.csv"},
   )


def _enrich(requirement_obj, db: Session) -> RequirementResponse:
    """将 ORM 对象转为 Response，并注入计算的 req_status"""
    resp = RequirementResponse.model_validate(requirement_obj)
    resp.req_status = attach_req_status(requirement_obj, db)
    return resp


@router.get("/", response_model=RequirementList)
def list_requirements(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    priority: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    city: Optional[List[str]] = Query(None, description="市（多选）"),
    district: Optional[str] = Query(None, description="区/县"),
    order_type: Optional[List[str]] = Query(None, description="工单类型（多选）"),
    date_from: Optional[str] = Query(None, description="接收日期起"),
    date_to: Optional[str] = Query(None, description="接收日期止"),
    db: Session = Depends(get_db),
):
    logger.info(f"list_requirements: city={city!r}, order_type={order_type!r}")
    items, total = RequirementService.list_requirements(
        db,
        skip=skip,
        limit=limit,
        priority=priority,
        status=status,
        search=search,
        city=city,
        district=district,
        order_type=order_type,
        date_from=date_from,
        date_to=date_to,
    )
    logger.info(f"result: total={total}, items={len(items)}")
    return RequirementList(
        total=total,
        items=[_enrich(r, db) for r in items],
        skip=skip,
        limit=limit,
    )


@router.post("/", response_model=RequirementResponse, status_code=201)
def create_requirement(
    data: RequirementCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    req = RequirementService.create_requirement(db, data)
    return _enrich(req, db)


@router.get("/{requirement_id}", response_model=RequirementResponse)
def get_requirement(requirement_id: int, db: Session = Depends(get_db)):
    req = RequirementService.get_requirement(db, requirement_id)
    if not req:
        raise HTTPException(status_code=404, detail="需求不存在")
    return _enrich(req, db)


@router.put("/{requirement_id}", response_model=RequirementResponse)
def update_requirement(
    requirement_id: int,
    data: RequirementUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    req = RequirementService.update_requirement(db, requirement_id, data)
    if not req:
        raise HTTPException(status_code=404, detail="需求不存在")
    return _enrich(req, db)


@router.delete("/{requirement_id}", status_code=204)
def delete_requirement(
    requirement_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    ok = RequirementService.delete_requirement(db, requirement_id)
    if not ok:
        raise HTTPException(status_code=404, detail="需求不存在")
    return None
