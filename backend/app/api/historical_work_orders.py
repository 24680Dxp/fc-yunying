import csv
import io
import os
import tempfile
from io import StringIO
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.schemas.historical_work_order import (
    HistoricalWorkOrderCreate,
    HistoricalWorkOrderList,
    HistoricalWorkOrderResponse,
    HistoricalWorkOrderUpdate,
)
from app.services.historical_work_order_service import HistoricalWorkOrderService
from app.auth import require_admin
from app.models.user import User

router = APIRouter(prefix="/api/v1/historical-work-orders", tags=["历史工单管理"])

FIELD_LABELS = {
    "site_code": "站点编号",
    "batch": "批次",
    "city": "地市",
    "access_address": "接入地址",
    "crm_order_id": "CRM订单号",
    "line_group_product_number": "专线集团产品号",
    "internet_work_order": "互联网工单",
    "internet_work_order_status": "互联网工单状态",
    "ql_crm_order_id": "千里眼CRM订单号",
    "ql_group_product_number": "千里眼集团产品号",
    "ql_work_order": "千里眼工单",
    "ql_work_order_status": "千里眼工单状态",
}

FIELDS_ORDER = list(FIELD_LABELS.keys())


@router.get("/", response_model=HistoricalWorkOrderList)
def list_work_orders(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    items, total = HistoricalWorkOrderService.list_work_orders(
        db, skip=skip, limit=limit, search=search, city=city,
    )
    return HistoricalWorkOrderList(
        total=total,
        items=[HistoricalWorkOrderResponse.model_validate(r) for r in items],
        skip=skip, limit=limit,
    )


@router.get("/export")
def export_work_orders(
    search: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    items, _ = HistoricalWorkOrderService.list_work_orders(
        db, skip=0, limit=100000, search=search, city=city,
    )
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([FIELD_LABELS[f] for f in FIELDS_ORDER])
    for item in items:
        writer.writerow([getattr(item, f, "") or "" for f in FIELDS_ORDER])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=historical_work_orders.csv"},
    )


@router.get("/cities", response_model=List[str])
def get_cities(db: Session = Depends(get_db)):
    return HistoricalWorkOrderService.get_cities(db)


@router.post("/", response_model=HistoricalWorkOrderResponse, status_code=201)
def create_work_order(
    data: HistoricalWorkOrderCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    wo = HistoricalWorkOrderService.create(db, data)
    return HistoricalWorkOrderResponse.model_validate(wo)


@router.put("/{work_order_id}", response_model=HistoricalWorkOrderResponse)
def update_work_order(
    work_order_id: int,
    data: HistoricalWorkOrderUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    wo = HistoricalWorkOrderService.update(db, work_order_id, data)
    if not wo:
        raise HTTPException(status_code=404, detail="历史工单不存在")
    return HistoricalWorkOrderResponse.model_validate(wo)


@router.delete("/{work_order_id}", status_code=204)
def delete_work_order(
    work_order_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    ok = HistoricalWorkOrderService.delete(db, work_order_id)
    if not ok:
        raise HTTPException(status_code=404, detail="历史工单不存在")
    return None


@router.post("/upload", status_code=200)
def upload_work_orders(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """上传 xlsx/csv 清单。站点编号唯一校验。"""
    fn = (file.filename or "").lower()
    if not fn.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .csv 文件")

    is_csv = fn.endswith(".csv")
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    if is_csv:
        text = content.decode("utf-8-sig")
        reader = csv.reader(StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=400, detail="CSV 文件为空")
        headers = all_rows[0]
        rows = all_rows[1:]
    else:
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            os.unlink(tmp_path)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"无法解析Excel文件: {str(e)}")
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Excel 文件没有工作表")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if not rows or not headers:
        raise HTTPException(status_code=400, detail="文件无数据")

    col_map = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        h_clean = str(h).strip()
        for field, label in FIELD_LABELS.items():
            if h_clean == label:
                col_map[idx] = field
                break

    if not col_map:
        raise HTTPException(
            status_code=400,
            detail="表头不匹配，请确保列名包含：" + "、".join(FIELD_LABELS.values()),
        )

    all_items = []
    for row in rows:
        if all(v is None for v in row):
            continue
        item = {}
        for col_idx, field in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                if hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                else:
                    val = str(val).strip()
            item[field] = val
        all_items.append(item)

    if not all_items:
        raise HTTPException(status_code=400, detail="未找到有效数据")

    result = HistoricalWorkOrderService.batch_import(db, all_items)
    return {
        "message": f"导入完成：成功 {result['imported']} 条，跳过 {result['skipped']} 条重复",
        **result,
    }
